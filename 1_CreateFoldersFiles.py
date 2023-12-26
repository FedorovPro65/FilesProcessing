# importing module
import os
from pathlib import Path
import openpyxl
import gzip

def mkdir_(Path_Folder: str):
    Path(Path_Folder).mkdir(parents=True, exist_ok=True)


def mkfile_(FileFullPath: str):
    with open(FileFullPath, 'w') as fp:
        fp.write(FileFullPath)


# Path_Folder = 'P:\\TaskFiles\\PROCESS_CSV\\A'
# mkdir_(Path_Folder)
# file_name = 'myfile.txt'
# FileFullPath = Path_Folder + '\\' + file_name
# mkfile_(FileFullPath)

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("File_List1.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
Path_Folder = ''
ar = {}
max_row = worksheet.max_row
for i in range(1, max_row):  # Начиная со строки с индексом 1 (вторая строка)
    j = 0
    rowx = []  # список для значений из строки файла
    for col in worksheet.iter_cols(1, worksheet.max_column):  # перебираем столбцы в строке
        j += 1
        print(col[i].value, end="; ")
        rowx.append(col[i].value)  # заполняем список значений

    print('')
    ar.setdefault(i, rowx)

print(ar)
Path_Folder = ''
for x, y in ar.items():  # перебираем значения из словаря, читаем названия папок и файлов, которые нужно создать.
    if Path_Folder != y[0]:  # если встретили новое название папки, то создаем ее.
        Path_Folder = y[0]
        mkdir_(Path_Folder)
    print(x, Path_Folder, y[0] + y[1], type(y[1]))
    if y[1][-3:]=='zip':
        FileFullPath = Path_Folder + '\\' + y[1][:-3] + 'csv'
    else:
        FileFullPath = Path_Folder + '\\' + y[1][:-3]


    print(FileFullPath)
    mkfile_(FileFullPath)  # Создаем файл по сформированному полному его пути

    FileFullPathGz = FileFullPath + '.gz'
    # FileFullPathGz.replace('\\', '/' )
    print(FileFullPathGz)
    with open(FileFullPath, 'rb') as f_in, gzip.open(FileFullPathGz, 'wb') as f_out:
        f_out.writelines(f_in)

    os.remove(FileFullPath)
# ---------------------------------
Path_Folder = 'P:\TaskFiles\ZIP\Rus'
os.chdir(Path_Folder)
print(Path_Folder)
directory = Path_Folder #os.fsencode(Path_Folder)
for file in os.listdir(directory):
    filename = os.fsdecode(file)
    if filename.endswith(".zip"):
        print(os.path.join(directory, filename))
        continue
    else:
        continue